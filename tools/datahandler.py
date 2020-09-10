"""
主要存放数据出路的方法
"""
import os
from openpyxl import load_workbook,Workbook
from openpyxl.worksheet.worksheet import Worksheet
import csv

class FileHanlder:
    def __init__(self):
        self.root_dir = os.path.join(os.path.dirname(__file__),'../')

    def write_excel(self,filepath,sheetname, row,col,value):
        filepath = os.path.join(self.root_dir,filepath)
        wb = load_workbook(filepath)
        print(wb.sheetnames)
        ws:Worksheet = wb[sheetname]
        ws.cell(row,col,value)
        wb.save(filepath)

    def read_excel(self,filepath,sheetname):
        excelpath = os.path.join(self.root_dir,filepath)
        workbook = load_workbook(excelpath)
        worksheet:Worksheet = workbook[sheetname]
        data = []
        for row in worksheet.iter_rows(min_row=1,max_row=worksheet.max_row,
                                       min_col=1,max_col=worksheet.max_column,
                                       values_only=True):
            data.append(row)
        return data
    def write_csv(self,filepath,filedata):
        pass

    def read_csv(self,filepath):
        csv_file = os.path.join(self.root_dir,filepath)
        with open(file=csv_file,mode='r',encoding='utf8') as csvfile:
            lines = csv.reader(csvfile)
            return [line for line in lines]

if __name__ == '__main__':
    fl = FileHanlder()
    data = fl.read_csv('data/login_data.csv')
    print(data)
    # data = fl.read_excel('data/data.xlsx','主题首页')
    # print(data)
    # fl.write_excel('data/data.xlsx','主题首页', 2,2,'helloworld')